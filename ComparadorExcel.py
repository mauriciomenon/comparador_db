# Python program to compare two excel files
#Autor: Rafael Henrique da Rosa
 
#O QUE DEFINE O PONTO SÃO OS DOIS PRIMEIROS CAMPOS



#TO DO:
    #comentar toda a parte da seleção de colunas
    #fazer a busca por nome da coluna ao inves do indice
    #verificar se marcar nenhum campo nas 3 seleçoes

import openpyxl,math,os,sys
from openpyxl.styles import PatternFill,Alignment,Font
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side
import tkinter as tk
from tkinter import ttk,Label,messagebox,filedialog as fd

str1=""
str2=""
str3="" 

######################################################################################################################################################
#PARTE DE COMPARAÇÃO
      
def compara(path_1, path_2):
    #vetores com o numero das colunas de comparação
    col_comp_index = {}
    #numeros de colunas para a comparação
    n_compara=0
    
    #transforma string em integer e adiciona ao vetor col_comp_index
    #o caracter 'e' vem do corte da palavra "nenhum", que ocorre quando nenhum campo é escolhido
    
    if(str1[1]!= "e"):
        col_comp_index[n_compara] = int(str1[1])
        n_compara+=1
    else:
        col_comp_index[n_compara] = 0
    if(str2[1]!= "e"):
        col_comp_index[n_compara] = int(str2[1])
        n_compara+=1
    else:
        col_comp_index[n_compara] = 0
    if(str3[1]!= "e"):
        col_comp_index[n_compara] = int(str3[1])
        n_compara+=1
    else:
        col_comp_index[n_compara] = 0

    
    #Se não foram selecionados arquivos não inicia a parte da comparação
    if(path_1 == "" or path_2 == ""):
        exit() #Esse exit funciona porem gera um erro
    
    #define borda e cores para preenchimento de celulas
    borda_fina = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    vermelho = PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')
    cinza = PatternFill(start_color='787878', end_color='787878', fill_type='solid')
    vermelho_claro = PatternFill(start_color='fa9696', end_color='fa9696', fill_type='solid')
    verde = PatternFill(start_color='82e89d', end_color='82e89d', fill_type='solid')
    
   
    #Tira o caminho e deixa só o nome dos arquivos selecionados
    while(path_1.find("/") != -1):
       path_1 = path_1[1:]  
    while(path_2.find("/") != -1):
        path_2 = path_2[1:]
    
    #Cria um Workbook de output, cria dois sheets e  muda seus nomes para o nome dos arquivos selecionados
    wb = openpyxl.Workbook()
    sheet_b1 = wb.active
    sheet_b2 = wb.create_sheet()
    sheet_b1.title = path_1
    sheet_b2.title = path_2
    
    #Copia célula por celula dos arquivos selecionados para o arqvuivo de output
    #O Openyxl não possui metodo para copiar o sheet inteiro de uma vez
    for row in banco1_sheet_obj:
        for cell in row:
            sheet_b1[cell.coordinate].value = cell.value
    for row in banco2_sheet_obj:
        for cell in row:
            sheet_b2[cell.coordinate].value = cell.value
    
    #cria um terceiro sheet no arquivo de output de relatório das ocorrencias        
    sheet_resul = wb.create_sheet()
    sheet_resul.title = "RELATÓRIO"
    
    #Deixa a primeira linha dos 3 sheets de output congelada para melhorar a visualização
    sheet_b2.freeze_panes = sheet_b2['A2']
    sheet_b1.freeze_panes = sheet_b2['A2']       
    sheet_resul.freeze_panes = sheet_b2['A2']       
    
    
    #Pinta a primeira linha do sheet 'banco1' do output de cinza, alinha e deixa em negrito
    for rows in sheet_b1.iter_rows(min_row=1, max_row=1, min_col=1):
        for cell in rows:
          cell.fill = cinza
          cell.alignment = Alignment(horizontal='center')
          cell.font = Font(bold= True)
          
    #Pinta a primeira linha do sheet 'banco2' do output de cinza, alinha e deixa em negrito
    for rows in sheet_b2.iter_rows(min_row=1, max_row=1, min_col=1):
        for i in range(sheet_b2.max_column):
          sheet_b2.cell(1,i+1).fill = cinza
          sheet_b2.cell(1,i+1).alignment = Alignment(horizontal='center')
          if(sheet_b1.cell(1,i+2).value == sheet_b2.cell(1,i+2).value):
              sheet_resul.cell(1,i+2).value = sheet_b1.cell(1,i+2).value
         # else:
              #Se o nome das colunas não sao iguais printa um erro e para o programa
          #    print("NOME DAS COLUNAS NÃO SÃO IGUAIS")
           #   sys.exit(0)
          sheet_b2.cell(1,i+1).font = Font(bold= True)
          
    #Pinta a primeira linha do sheet 'Relatório' do output de cinza, alinha e deixa em negrito
    #i+1 desloca as colunas 1 coluna para a direita para anexar uma coluna index no inicio do relatório
    #a coluna index mostra o numero da linha que foi copiada para o relatório      
    for rows in sheet_resul.iter_rows(min_row=1, max_row=1, min_col=1):
        sheet_resul.cell(1,1).value = "INDEX"
        for i in range(sheet_resul.max_column):
          sheet_resul.cell(1,i+1).fill = cinza
          sheet_resul.cell(1,i+1).alignment = Alignment(horizontal='center')
          sheet_resul.cell(1,i+1).font = Font(bold= True)
     
    
    #Ajusta o comprimento das colunas do sheet 'banco1'    
    dims = {}
    for row in sheet_b1.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))    
       
    for i, column_width in dims.items():  # ,1 to start at 1
        sheet_b1.column_dimensions[get_column_letter(i)].width = int(math.ceil(column_width*1.42))
    
    #Ajusta o comprimento das colunas do sheet 'banco2' e deixa o resultado igual o banco2        
    for row in sheet_b2.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))    
    
    for i, column_width in dims.items():
        sheet_b2.column_dimensions[get_column_letter(i)].width = int(math.ceil(column_width*1.42))
        sheet_resul.column_dimensions[get_column_letter(i+1)].width = int(math.ceil(column_width*1.42))
        
        
    #################################################################################################################################################
    #Parte do resultado
    
    #Array para armazenar as linhas iguais encontradas e as discr   epancias    
    encontrado = {}
    discrep = {}
    
    #Zera os arrays anteriores (Poderia utilizar outra lógica para evitar)
    for i in range(sheet_b1.max_row):
        encontrado[i] = 0
        discrep[i] = 0
    
    corresp = {}
    for i in range(sheet_b2.max_column):
        corresp[i] = 0    
    for i in range(1,sheet_b2.max_column):
        corresp[i] = 0
        for j in range(1,sheet_b1.max_column):
            if(sheet_b1.cell(1,i).value == sheet_b2.cell(1,j).value):
                corresp[i] = j
        
    #k representa a linha que ta sendo escrita no sheet de relatorio
    k=3
    sheet_resul.cell(k,1).value = "LINHAS EXCLUIDAS (presentes no banco 1 e não no banco 2):"
    k+=1
    
    #Linhas excluidas (presentes no banco 1 e não no banco 2):

    #Percorre cada linha nos sheets banco1 e banco2       
    for i in range(sheet_b1.max_row):
        for j in range(sheet_b2.max_row):
            flag_1 = 0
            #faz uma iteração em todos os indices que serão comparados
            for m in range(n_compara):
                #se o campo marcado é diferente de "nenhum"
                if(col_comp_index[m]!=0):
                    #para cada indice, compara as celulas no sheet_b1 e sheet_b2 adicionando 1 a uma flag
                    if(sheet_b1.cell(i+1,col_comp_index[m]).value == sheet_b2.cell(j+1,corresp[col_comp_index[m]]).value):
                        flag_1+=1
            #caso todas as colunas a serem comparadas sejam iguais nos dois sheets
            if(flag_1 == n_compara):
                #marca a linha como encontrada nos dois bancos
                encontrado[i] = j+1
                #checa se existem discrepancias percorrendo cada coluna da linha
                for l in range(sheet_b1.max_column -1):
                    if(sheet_b1.cell(i+1,l+1).value != sheet_b2.cell(j+1,corresp[l+1]).value):
                        discrep[i] = j
                        #muda a cor das discrepancias
                        sheet_b1.cell(i+1,l+1).fill = vermelho
                        sheet_b2.cell(j+1,l+1).fill = vermelho
                        #pinta as celulas das colunas comparadas
                        for n in range(n_compara):
                            sheet_b1.cell(i+1,col_comp_index[n]).fill = vermelho
                            sheet_b2.cell(j+1,col_comp_index[n]).fill = vermelho
                        
                break
        #Se a linha não for encontrada    
        if(encontrado[i] == 0):
           #Marca a linha como uma linha excluida no sheet banco1 e copia para o sheet de relatório
           for l in range(sheet_b1.max_column-1):
               sheet_b1.cell(i+1,l+1).fill = vermelho_claro
               sheet_b1.cell(i+1,l+1).fill = vermelho_claro
               sheet_resul.cell(k,l+1+1).value = sheet_b1.cell(i+1,l+1).value
               sheet_resul.cell(k,l+1).fill = vermelho_claro
               sheet_resul.cell(k,l+1).border = borda_fina
               sheet_resul.cell(k, 1).value = i+1
           sheet_resul.cell(k, 1).font = Font(bold= True)
           sheet_resul.cell(k, 1).fill = cinza 
           k+=1
           
    #Linhas discrepantes:
        
    k+=2
    sheet_resul.cell(k,1).value = "LINHAS DISCREPANTES:"
    k+=1
    for i in range(sheet_b1.max_row):
        if(discrep[i] !=0):
            sheet_resul.cell(k,1).value = path_1
            sheet_resul.cell(k+2,1).value = path_2
            k+=1
            sheet_resul.cell(k, 1).value = i+1
            sheet_resul.cell(k, 1).font = Font(bold= True)
            sheet_resul.cell(k, 1).fill = cinza
            sheet_resul.cell(k+2, 1).value = discrep[i]+1
            sheet_resul.cell(k+2, 1).font = Font(bold= True)
            sheet_resul.cell(k+2, 1).fill = cinza
            sheet_resul.cell(k,l+1+1).fill = vermelho
            for l in range(sheet_b1.max_column-1):
               sheet_resul.cell(k,l+1+1).value = sheet_b1.cell(i+1,l+1).value
               sheet_resul.cell(k,l+1).border = borda_fina
               sheet_resul.cell(k+2,l+1+1).value = sheet_b2.cell(discrep[i]+1,l+1).value
               sheet_resul.cell(k+2,l+1).border = borda_fina
               if(sheet_resul.cell(k,l+1+1).value != sheet_resul.cell(k+2,l+1+1).value):
                   sheet_resul.cell(k,l+1+1).fill = vermelho
                   sheet_resul.cell(k+2,l+1+1).fill = vermelho
            k +=4
    k +=3
    
    
    #Linhas Novas (presentes no banco 2 e não no banco 1):
    
    sheet_resul.cell(k,1).value = "LINHAS NOVAS (Presentes somente no Banco2):" 
    k +=1
    for i in range(sheet_b2.max_row):
        flag = 0
        for j in range(sheet_b1.max_row):
            if(i+1 == encontrado[j]):
                flag = 1
                break
        if(flag==0):
            for l in range(sheet_b2.max_column-1):
                sheet_resul.cell(k,l+1+1).value = sheet_b2.cell(i+1,l+1).value
                sheet_resul.cell(k,l+1).border = borda_fina
                sheet_resul.cell(k,l+1).fill = verde
                sheet_b2.cell(i+1,l+1).fill = verde
                sheet_b2.cell(i+1,l+1).fill = verde
            sheet_resul.cell(k, 1).value = i+1
            sheet_resul.cell(k, 1).font = Font(bold= True)
            sheet_resul.cell(k, 1).fill = cinza
            k +=1
    
    
    #Retira o grid do sheet de resultados        
    sheet_resul.sheet_view.showGridLines = False
    #salva o arquivo output na mesma pasta do arquivo py e abre
    wb.save("output.xlsx")
    os.system("start EXCEL.EXE output.xlsx")
   # sys.exit() 
    #https://stackoverflow.com/questions/13197574/openpyxl-adjust-column-width-size
    
##################################################################################################################################################
#PARTE TK


path_1 = ""
path_2 = ""
#cria a janela principal, que não pode ser redimensionada 
root = tk.Tk()
root.title('Comparador Excel (.xlsx) V0.1.2')
root.resizable(False, False)
root.geometry('400x350')


#valores padrão dos campos de seleção da comparação
str1 = "3"
str2 = "4"
str3 = "e"
   
#função acionada quando o botão iniciar é pressionado
def clicked():
    if(path_1 != "" and path_2 != ""):
        #tenta destruir a janela de seleção de comparação caso ela esteja aberta
        try:
            selecao_col.destroy()
        except:
            pass
        #inicia a comparação e destroi a janela principal 
        compara(path_1,path_2)
        root.destroy()
    else:
        messagebox.showinfo("ERRO","Selecione os arquivos")
        
#função acionada quando o botão de convirmar a seleção dos campos de comparação é pressionado        
def confirma_selecao():
  #  selecao_col.destroy()
    if(path_1 != "" and path_2 != ""):
        #destroi a janela de seleção
        selecao_col.destroy()
        #exibe a janela principal
        root.deiconify()
        
        #essas variáveis são globais pois são utilizadas em outras funções e na função principal de comparação
        #TO DO: transformar em variáveis locais que são passadas a outras funções
        global variable
        global variable_2
        global variable_3
    
        global str1
        global str2
        global str3    
        
        #Coleta os valores dos campos de seleção
        tmp = variable.get()
        str1 = tmp
        tmp = variable_2.get()
        str2 = tmp
        tmp = variable_3.get()
        str3 = tmp
    else:
        messagebox.showinfo("ERRO","Selecione os arquivos")

#função acionada quando o botão  seleção dos campos de comparação é pressionado        
def colselect():
    if(path_1 == "" or path_2 == ""):
       messagebox.showinfo("ERRO","Selecione os arquivos")
    else:
        #essas variáveis são globais pois são utilizadas em outras funções e na função principal de comparação
        #TODO: transformar em variáveis locais que são passadas a outras funções
        global variable
        global variable_2
        global variable_3
        
        global str1
        global str2
        global str3 
        
        #valores padrão

        
        #esconde a janela principal
        root.withdraw()
        global selecao_col
        #inicia a janela de seleção
        selecao_col = tk.Tk()
        
        #informações da janela de seleção
        selecao_col.title('Seleção de colunas para comparação')
        selecao_col.resizable(False, False)
        selecao_col.geometry('400x350')
        
        #array com o nome das colunas
        header = []
        #adiciona a string "Nenhum no inicio do vetor para esta ser uma opção de seleção"
        header.append("Nenhum")
        
        #Percorre as colunas do sheet_b1 guardando no array header o indice e nome de cada coluna
        for i in range(banco1_sheet_obj.max_column):
            if(banco1_sheet_obj.cell(1,i+1).value == banco2_sheet_obj.cell(1,i+1).value):
                str_tmp =''.join(banco1_sheet_obj.cell(1,i+1).value)
                str_tmp = "[" + (str)(i+1) + "] " + str_tmp
              #  print(str_tmp)
                header.append(str_tmp)
         #   else:
                #TODO: comparar as tabelas mesmo se tiverem colunas diferentes 
         #       print("tabelas com colunas diferentes")    

        #configuração da janela de seleção com 3 campos de seleção
        variable = tk.StringVar(selecao_col)
        variable.set(str1)# default value
        variable_2 = tk.StringVar(selecao_col)
        variable_2.set(str2) # default value
        variable_3 = tk.StringVar(selecao_col)
        variable_3.set(str3) # default value
        selecao_campo_1 = tk.OptionMenu(selecao_col, variable,*header)
        selecao_campo_1.place(x=100, y = 80,height = 30, width = 200)
        selecao_campo_2 = tk.OptionMenu(selecao_col, variable_2,*header)
        selecao_campo_2.place(x=100, y = 130,height = 30, width = 200)
        selecao_campo_3 = tk.OptionMenu(selecao_col, variable_3,*header)
        selecao_campo_3.place(x=100, y = 180,height = 30, width = 200)
        lbl5 = Label(selecao_col, text="")
        lbl5.configure(text="Selecione as colunas que serão utilizados de base para comparação:")
        lbl5.place(x=0, y =10,height = 50, width = 400)      
        botao_confirma_coluna = ttk.Button(selecao_col,text='Salvar',command=confirma_selecao)
        botao_confirma_coluna.place(x=150, y = 230,height = 30, width = 100)

        #Coleta os valores dos campos de seleção   
        tmp = variable.get()
        str1 = tmp
        tmp = variable_2.get()
        str2 = tmp
        tmp = variable_3.get()
        str3 = tmp
        selecao_col.protocol("WM_DELETE_WINDOW", close_colsel)
        selecao_col.mainloop()


#função acionada quando o botão '?' é pressionado        
def myinfo():
    #mostra as informações do algoritmo
    messagebox.showinfo("Info","Autor: Rafael Henrique da Rosa\nEstagiário Itaipu Binacional - SMIN.DT - Março de 2022\n\
O algoritmo compara duas tabelas e mostra em uma planilha de resultado as linhas excluidas,novas e discrepantes")


#função acionada quando o botão 'selecionar banco antigo' é pressionado        
def select_file():
    file_types = (('Excel Files', '*.xlsx'),('All files', '*.*'))
    file_name = fd.askopenfilename(title='Selecionar Banco antigo',filetypes=file_types)
    
    #essas variáveis são globais pois os woorkbooks são abertos aqui para pegar o nome das colunas
        #porem são acessados na função de comparação
    global path_1
    global banco1_obj
    global banco1_sheet_obj
    
    #str para manipulação do caminho
    path_1 = file_name
    
    #tenta abrir o primeiro woorkbook de duas maneiras diferentes
    try:
       
        banco1_obj = openpyxl.load_workbook(path_1)
    except:
        try:
            path_1 = path_1.replace("/", "\\")
            banco1_obj = openpyxl.load_workbook(path_1)
        except:
            print("ai realmente não ta abrindo o arquivo antigo")
    
    #carrega o sheet        
    banco1_sheet_obj = banco1_obj.active
    
    #Se não funcionar descomentar as 3 linhas da sequencia
    #global str1
    #global str2
    #global str3
 
    #manipulação do caminho do arquivo para label da janela principal
    path_1_ = path_1
    while(path_1_.find("/") != -1):
        path_1_ = path_1_[1:]
    lbl1.configure(text=path_1_)
    global variable
    global variable_2
    global variable_3
       
    global str1
    global str2
    global str3 
    str1 = "[3] "+ banco1_sheet_obj.cell(1,3).value
    str2 = "[4] "+ banco1_sheet_obj.cell(1,4).value
    str3 = "Nenhum"

#função acionada quando o botão 'selecionar banco novo' é pressionado            
def select_file2():
    file_types = (('Excel Files', '*.xlsx'), ('All files', '*.*') )
    file_name = fd.askopenfilename(title='Selecionar Banco novo',filetypes=file_types)
    
    #essas variáveis são globais pois os woorkbooks são abertos aqui para pegar o nome das colunas
        #porem são acessados na função de comparação
    global path_2
    global banco2_obj
    global banco2_sheet_obj
    
    #str para manipulação do caminho
    path_2 = file_name
    
    #tenta abrir o segundo woorkbook de duas maneiras diferentes
    try:
        banco2_obj = openpyxl.load_workbook(path_2)
    except:
        try:
            path_2 = path_2.replace("/", "\\")
            banco2_obj = openpyxl.load_workbook(path_2)
        except:
            print("ai realmente não ta abrindo o arquivo novo")

    #carrega o sheet        
    banco2_sheet_obj = banco2_obj.active
    
    #manipulação do caminho do arquivo para label da janela principal
    path_2_ = path_2
    while(path_2_.find("/") != -1):
       path_2_ = path_2_[1:]
    lbl2.configure(text=path_2_)


#labels e botoes da janela principal    
lbl4 = Label(root, text="")
lbl4.configure(text="*A comparação pode demorar de acordo com o tamanho do banco")
lbl4.place(x=0, y =250,height = 50, width = 400)

botao_banco_antigo = ttk.Button(root,text='Selecionar Banco antigo',command=select_file)
botao_banco_antigo.place(x=100, y = 10,height = 40, width = 200)
lbl1 = Label(root, text="")
lbl1.configure(text="")
lbl1.place(x=100, y =50,height = 30, width = 200)

botao_banco_novo = ttk.Button(root,text='Selecionar Banco novo',command=select_file2)
botao_banco_novo.place(x=100, y = 90,height = 40, width = 200)
lbl2 = Label(root, text="")
lbl2.configure(text="")
lbl2.place(x=100, y = 130,height = 30, width = 200)

botao_iniciar = ttk.Button(root,text='Iniciar',command=clicked)
botao_iniciar.place(x=150, y = 180,height = 30, width = 100)
lbl3 = Label(root, text="")
lbl3.configure(text="") 
lbl3.place(x=100, y = 210,height = 30, width = 200)

botao_info = ttk.Button(root,text='?',command=myinfo)
botao_info.place(x=350, y = 300,height = 32, width = 32)

botao_selecao = ttk.Button(root,text='Colunas de comparação',command=colselect)
botao_selecao.place(x=50, y = 300,height = 32, width = 150)


# run the application


def close_root():
    if messagebox.askokcancel("SAIR", "Deseja Sair?"):
        root.destroy()
        sys.exit()
def close_colsel():
    root.deiconify()
    selecao_col.destroy()
    

root.protocol("WM_DELETE_WINDOW", close_root)

root.mainloop()
